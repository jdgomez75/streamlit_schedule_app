import streamlit as st
import pandas as pd
import numpy as np
import streamlit_authenticator as stauth
import bcrypt
import yaml
from yaml.loader import SafeLoader
from src.database import Database
from datetime import datetime, timedelta
from io import BytesIO
from decimal import Decimal

def convert_to_native(obj):
    """Convierte CUALQUIER tipo a tipos nativos Python"""
    if obj is None:
        return None
    if isinstance(obj, np.ndarray):
        if obj.size == 1:
            return convert_to_native(obj.item())
        else:
            return obj.tolist()
    if isinstance(obj, np.integer):
        return int(obj)
    if isinstance(obj, np.floating):
        return float(obj)
    if isinstance(obj, np.bool_):
        return bool(obj)
    if isinstance(obj, Decimal):
        return float(obj)
    if isinstance(obj, (str, int, float, bool)):
        return obj
    if isinstance(obj, (list, tuple)):
        return [convert_to_native(item) for item in obj]
    if isinstance(obj, dict):
        return {k: convert_to_native(v) for k, v in obj.items()}
    return str(obj)

def _row_to_dict(cursor, row):
    """Convierte una fila de PostgreSQL a diccionario, manejando tipos problemáticos."""
    if isinstance(row, dict):
        return row
        
    columns = [desc[0] for desc in cursor.description]
    data_dict = dict(zip(columns, row))
    
    # NUEVA LÓGICA: Iterar y convertir tipos problemáticos
    for key, value in data_dict.items():
        if isinstance(value, Decimal):
            # Convierte Decimal de DB a float nativo de Python
            data_dict[key] = float(value)
        elif isinstance(value, (date, time)):
            # Convierte objetos date/time de DB a string (para evitar conflictos de NumPy)
            data_dict[key] = str(value)
        # Puedes añadir más conversiones aquí si es necesario
        
    return data_dict


# Configuración de la página
st.set_page_config(
    page_title="Admin - Rubí Mata Salón",
    page_icon="👑",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personalizado para admin
st.markdown("""
<style>
    /* Estilo general */
    .main {
        background: linear-gradient(135deg, #F0F9FF 0%, #E0E7FF 100%);
    }
    
    /* Cards de estadísticas */
    .stat-card {
        background: white;
        padding: 1.5rem;
        border-radius: 15px;
        box-shadow: 0 4px 6px rgba(0,0,0,0.1);
        text-align: center;
    }
    
    .stat-number {
        font-size: 2.5rem;
        font-weight: bold;
        color: #3B82F6;
        margin: 0;
    }
    
    .stat-label {
        color: #6B7280;
        font-size: 0.9rem;
        margin-top: 0.5rem;
    }
    
    /* Calendario tipo agenda */
    .appointment-card {
        background: white;
        padding: 1rem;
        border-radius: 12px;
        margin-bottom: 0.5rem;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        border-left: 4px solid #3B82F6;
        transition: all 0.3s ease;
    }
    
    .appointment-card:hover {
        transform: translateX(5px);
        box-shadow: 0 4px 8px rgba(59,130,246,0.3);
    }
    
    .appointment-card.status-pending {
        border-left-color: #F59E0B;
    }
    
    .appointment-card.status-confirmed {
        border-left-color: #10B981;
    }
    
    .appointment-card.status-cancelled {
        border-left-color: #EF4444;
    }
    
    /* Timeline view */
    .timeline-container {
        position: relative;
        padding-left: 100px;
    }
    
    .timeline-hour {
        position: absolute;
        left: 0;
        width: 80px;
        text-align: right;
        color: #6B7280;
        font-weight: 600;
    }
    
    .timeline-slot {
        min-height: 60px;
        border-top: 1px solid #E5E7EB;
        padding: 0.5rem 0;
    }
    
    /* Badges */
    .badge {
        display: inline-block;
        padding: 0.25rem 0.75rem;
        border-radius: 12px;
        font-size: 0.75rem;
        font-weight: 600;
    }
    
    .badge-pending {
        background: #FEF3C7;
        color: #92400E;
    }
    
    .badge-confirmed {
        background: #D1FAE5;
        color: #065F46;
    }
    
    .badge-paid {
        background: #DBEAFE;
        color: #1E40AF;
    }
    
    /* Ocultar elementos de Streamlit */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
</style>
""", unsafe_allow_html=True)

# Inicializar base de datos
@st.cache_resource
def init_db():
    return Database()

db = init_db()

# ==================== AUTENTICACIÓN (Adaptación forzada a versión MUY ANTIGUA) ====================
db = init_db()
import streamlit_authenticator as stauth # Asegúrate que el import esté aquí o al inicio

try:
    # 1. Obtener usuarios desde la base de datos
    usernames_list, hashed_passwords_list, names_list = db.get_all_users_for_auth()
    
    # 2. Ensamblar las credenciales en el formato de diccionario ANIDADO
    credentials = {
        'usernames': {}
    }
    
    for i, username in enumerate(usernames_list):
        credentials['usernames'][username] = {
            # Los nombres de los campos DEBEN coincidir con el código original
            'email': f'{username}@admin.app', 
            'name': names_list[i],
            'password': hashed_passwords_list[i]
        }

    # 3. Inicializar el autenticador pasándole SOLO el diccionario de credenciales como primer argumento.
    # En versiones muy antiguas, se acepta un único argumento posicional de diccionario.
    authenticator = stauth.Authenticate(
        credentials, # 1er argumento: el diccionario de credenciales
        'streamlit_auth', # 2do argumento: cookie_name
        'auth_key',       # 3er argumento: cookie_key
        30                # 4to argumento: cookie_expiry_days
    )
    
except Exception as e:
    st.error(f"Error al cargar usuarios para autenticación: {e}")
    st.info("Asegúrate de que la tabla 'users' exista y los métodos de DB estén correctos.")
    st.stop()
# ===========================================================================


# ==================== LÓGICA DE INICIO DE SESIÓN (Final) ====================

# Usar la sintaxis posicional simple.
# El error de Location desaparece con la versión 0.2.2
name, authentication_status, username = authenticator.login('Login', 'main') 

if authentication_status:
    # Usuario autenticado exitosamente
    st.session_state["authentication_status"] = authentication_status
    st.session_state["name"] = name
    st.session_state["username"] = username
    
elif authentication_status is False:
    # Credenciales incorrectas
    st.error('Nombre de usuario/contraseña incorrectos')
    st.stop()
    
elif authentication_status is None:
    # Esperando Login
    st.warning('Por favor, ingresa tu nombre de usuario y contraseña')
    st.stop()
# =========================================================================


# Funciones auxiliares
def get_status_badge(status):
    """Retorna badge HTML según el estado"""
    badges = {
        'pending': '<span class="badge badge-pending">⏳ Pendiente</span>',
        'confirmed': '<span class="badge badge-confirmed">✅ Confirmada</span>',
        'cancelled': '<span class="badge badge-pending">❌ Cancelada</span>',
        'completed': '<span class="badge badge-confirmed">✓ Completada</span>'
    }
    return badges.get(status, status)

def get_payment_badge(payment_status):
    """Retorna badge HTML según el estado de pago"""
    badges = {
        'pending': '<span class="badge badge-pending">💳 Anticipo pendiente</span>',
        'partial': '<span class="badge badge-paid">💰 Anticipo pagado</span>',
        'paid': '<span class="badge badge-confirmed">✅ Pagado completo</span>'
    }
    return badges.get(payment_status, payment_status)

def format_time_range(start, end):
    """Formatea rango de tiempo"""
    return f"{start} - {end}"

def calculate_stats(bookings):
    """Calcula estadísticas del día"""
    total_bookings = len(bookings)
    total_revenue = sum(b['total_price'] for b in bookings)
    deposits_collected = sum(b['deposit_paid'] for b in bookings)
    pending_payments = total_revenue - deposits_collected
    
    confirmed = len([b for b in bookings if b['status'] == 'confirmed'])
    pending = len([b for b in bookings if b['status'] == 'pending'])
    
    return {
        'total_bookings': total_bookings,
        'confirmed': confirmed,
        'pending': pending,
        'total_revenue': total_revenue,
        'deposits_collected': deposits_collected,
        'pending_payments': pending_payments
    }

def get_payment_status(total_price, deposit_paid):
    """Calcula el estado de pago basado en los montos"""
    if deposit_paid <= 0:
        return 'pending', '💳 Anticipo pendiente'
    elif deposit_paid >= total_price:
        return 'paid', '✅ Pagado completo'
    else:
        return 'partial', '💰 Anticipo pagado'

def get_payment_badge_from_amounts(total_price, deposit_paid):
    """Retorna badge HTML del estado de pago"""
    if deposit_paid <= 0:
        return '<span class="badge badge-pending">💳 Anticipo pendiente</span>'
    elif deposit_paid >= total_price:
        return '<span class="badge badge-confirmed">✅ Pagado completo</span>'
    else:
        return '<span class="badge badge-paid">💰 Anticipo pagado</span>'

# ============================================
# FUNCIÓN 1: Exportar Citas a Excel
# ============================================
#
# Agregar esta función DESPUÉS de las vistas (antes de main())

def export_bookings_to_excel():
    """
    Exporta citas del mes actual a Excel con información completa
    
    Estructura:
    - Obtiene bookings con relaciones (clientes, profesionales, servicios, pagos)
    - Cálcula monto pendiente (total - depósito pagado)
    - Formatea moneda, fechas y horas
    - Aplica estilos y colores por estado
    """
    try:
        from datetime import datetime
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        # Obtener citas del mes actual CON TODAS LAS RELACIONES
        with db.get_connection() as conn:
            cursor = conn.cursor()
            
            # QUERY COMPLETA - Une todas las relaciones
            cursor.execute('''
                SELECT 
                    b.booking_code as "Código Cita",
                    b.client_name as "Cliente",
                    b.client_phone as "Teléfono",
                    b.client_email as "Email",
                    b.date as "Fecha",
                    b.start_time as "Hora Inicio",
                    b.end_time as "Hora Fin",
                    COALESCE(p.name, 'Sin asignar') as "Profesional",
                    STRING_AGG(DISTINCT s.name, ', ') as "Servicios",
                    STRING_AGG(DISTINCT c.name, ', ') as "Categoría",
                    b.total_price as "Total",
                    b.deposit_paid as "Depósito Pagado",
                    (b.total_price - b.deposit_paid) as "Pendiente",
                    COALESCE(pay.amount, 0) as "Monto Pagado",
                    COALESCE(pay.payment_status, '-') as "Estado Pago",
                    b.status as "Estado Cita"
                FROM bookings b
                LEFT JOIN professionals p ON b.professional_id = p.id
                LEFT JOIN booking_services bs ON b.id = bs.booking_id
                LEFT JOIN services s ON bs.service_id = s.id
                LEFT JOIN categories c ON s.category_id = c.id
                LEFT JOIN payments pay ON b.booking_code = pay.booking_code
                WHERE DATE_TRUNC('month', b.date::timestamp) = DATE_TRUNC('month', CURRENT_DATE)
                GROUP BY b.id, p.id, p.name, pay.id, pay.amount, pay.payment_status
                ORDER BY b.date DESC, b.start_time ASC
            ''')
            
            columns = [desc[0] for desc in cursor.description]
            rows = cursor.fetchall()
        
        if not rows:
            st.warning("📭 No hay citas para exportar este mes")
            return None
        
        # Crear DataFrame
        df = pd.DataFrame(rows, columns=columns)
        
        # ========== FORMATEO DE DATOS ==========
        
        # Fecha: convertir a formato YYYY-MM-DD
        df['Fecha'] = pd.to_datetime(df['Fecha']).dt.strftime('%Y-%m-%d')
        
        # Horas: convertir a string HH:MM:SS
        df['Hora Inicio'] = df['Hora Inicio'].astype(str)
        df['Hora Fin'] = df['Hora Fin'].astype(str)
        
        # Moneda: formatear con $ y 2 decimales
        df['Total'] = df['Total'].astype(float).apply(lambda x: f"${x:,.2f}" if pd.notna(x) else "$0.00")
        df['Depósito Pagado'] = df['Depósito Pagado'].astype(float).apply(lambda x: f"${x:,.2f}" if pd.notna(x) else "$0.00")
        df['Pendiente'] = df['Pendiente'].astype(float).apply(lambda x: f"${x:,.2f}" if pd.notna(x) else "$0.00")
        df['Monto Pagado'] = df['Monto Pagado'].astype(float).apply(lambda x: f"${x:,.2f}" if pd.notna(x) else "$0.00")
        
        # Estado: capitalizar
        df['Estado Cita'] = df['Estado Cita'].str.capitalize()
        df['Estado Pago'] = df['Estado Pago'].str.capitalize()
        
        # Crear Excel en memoria
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Citas Mes', index=False)
            
            workbook = writer.book
            worksheet = writer.sheets['Citas Mes']
            
            # ===== ESTILOS =====
            
            # Bordes finos
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Color de encabezado (pink)
            header_fill = PatternFill(start_color="EC4899", end_color="EC4899", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Aplicar estilos al encabezado
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Altura del encabezado
            worksheet.row_dimensions[1].height = 30
            
            # ===== APLICAR FORMATO A DATOS =====
            
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                
                # Obtener estado de la cita (última columna)
                status_cell = row[-2]  # Estado Cita es penúltima columna
                status_value = str(status_cell.value).lower() if status_cell.value else ""
                
                # Color de fondo según estado
                if 'confirmed' in status_value:
                    row_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")  # Verde
                elif 'pending' in status_value:
                    row_fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")  # Amarillo
                elif 'cancelled' in status_value:
                    row_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # Rojo
                else:
                    row_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")  # Gris
                
                # Aplicar a cada celda
                for cell in row:
                    cell.fill = row_fill
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
            # ===== AJUSTAR ANCHO DE COLUMNAS =====
            
            column_widths = {
                'A': 18,  # Código Cita
                'B': 20,  # Cliente
                'C': 14,  # Teléfono
                'D': 25,  # Email
                'E': 12,  # Fecha
                'F': 12,  # Hora Inicio
                'G': 12,  # Hora Fin
                'H': 18,  # Profesional
                'I': 30,  # Servicios
                'J': 20,  # Categoría
                'K': 12,  # Total
                'L': 15,  # Depósito Pagado
                'M': 14,  # Pendiente
                'N': 14,  # Monto Pagado
                'O': 15,  # Estado Pago
                'P': 14,  # Estado Cita
            }
            
            for col_letter, width in column_widths.items():
                worksheet.column_dimensions[col_letter].width = width
            
            # ===== CONGELAR ENCABEZADO =====
            worksheet.freeze_panes = 'A2'
        
        output.seek(0)
        print(f"✅ Excel generado: {len(rows)} citas exportadas")
        return output.getvalue()
        
    except Exception as e:
        print(f"❌ Error exportando Excel: {e}")
        import traceback
        traceback.print_exc()
        st.error(f"Error al generar Excel: {str(e)}")
        return None




# ============================================
# FUNCIÓN 2: Enviar Recordatorios
# ============================================
#
# Agregar esta función DESPUÉS de export_bookings_to_excel()

def send_appointment_reminders():
    """
    Envía recordatorios de citas para MAÑANA
    
    Estructura de query:
    - bookings: date (DATE), start_time (TIME), end_time (TIME)
    - professionals: name (VARCHAR)
    - Filtra por: status NOT IN ('cancelled', 'completed')
    
    Retorna: cantidad de recordatorios enviados, o -1 si hay error
    """
    try:
        from src import notifications
        from datetime import datetime, timedelta
        
        tomorrow = datetime.now().date() + timedelta(days=1)
        
        print(f"📅 Enviando recordatorios para: {tomorrow}")
        
        # Obtener citas de mañana
        with db.get_connection() as conn:
            cursor = conn.cursor()
            
            # QUERY CORREGIDA
            cursor.execute('''
                SELECT 
                    b.id,
                    b.booking_code,
                    b.client_name,
                    b.client_email,
                    b.date,
                    b.start_time,
                    b.end_time,
                    COALESCE(p.name, 'Sin asignar') as professional_name
                FROM bookings b
                LEFT JOIN professionals p ON b.professional_id = p.id
                WHERE b.date = %s
                AND b.status NOT IN ('cancelled', 'completed')
                ORDER BY b.start_time ASC
            ''', (tomorrow,))
            
            bookings = cursor.fetchall()
            print(f"🔍 {len(bookings)} citas encontradas para mañana")
        
        if not bookings:
            print("ℹ️ No hay citas para mañana")
            return 0
        
        # Enviar recordatorio a cada cliente
        sent_count = 0
        failed_count = 0
        
        for booking in bookings:
            booking_id, code, name, email, date, start_time, end_time, prof_name = booking
            
            try:
                # Validar email
                if not email or not email.strip():
                    print(f"⚠️ Email vacío para cita {code}")
                    failed_count += 1
                    continue
                
                # Preparar datos para el email
                booking_data = {
                    'client': {
                        'name': name,
                        'email': email
                    },
                    'appointment': {
                        'date': str(date),
                        'start_time': str(start_time),
                        'end_time': str(end_time)
                    },
                    'booking_code': code,
                    'professional': {
                        'name': prof_name or 'Profesional'
                    }
                }
                
                # Enviar email
                success = notifications.enviar_recordatorio_cita(booking_data)
                
                if success:
                    sent_count += 1
                    print(f"✅ Recordatorio enviado a {email} (Cita: {code})")
                else:
                    failed_count += 1
                    print(f"⚠️ Error enviando a {email} (Cita: {code})")
                    
            except Exception as e:
                failed_count += 1
                print(f"⚠️ Excepción en {email}: {str(e)}")
        
        print(f"📊 Resultado: {sent_count} enviados, {failed_count} fallidos")
        return sent_count
        
    except Exception as e:
        print(f"❌ Error en send_appointment_reminders: {e}")
        import traceback
        traceback.print_exc()
        st.error(f"Error enviando recordatorios: {str(e)}")
        return -1

# ==================== SIDEBAR ====================

with st.sidebar:
    st.markdown("# 👑 Panel Admin")
    st.markdown("---")
    
    # Selector de vista
    view_mode = st.radio(
    "Vista",
    ["📅 Calendario del Día", "📊 Agenda Semanal", "💳 Gestión de Pagos", "📈 Reportes", "⚙️ Configuración"],
    index=0
)
    
    st.markdown("---")
    
    # Inicializar fecha en session_state si no existe
    if 'selected_date' not in st.session_state:
        st.session_state.selected_date = datetime.now().date()
    
    # Selector de fecha
    selected_date = st.date_input(
        "Fecha",
        value=st.session_state.selected_date,
        key="date_selector"
    )
    
    # Sincronizar el valor del widget con session_state
    if selected_date != st.session_state.selected_date:
        st.session_state.selected_date = selected_date
        st.rerun()
    
    st.markdown("---")
    
    # Filtros
    st.markdown("### 🔍 Filtros")
    
    # Obtener profesionales
    with db.get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id, name FROM professionals WHERE active = TRUE")
        professionals = [dict(zip([desc[0] for desc in cursor.description], row)) for row in cursor.fetchall()]
    
    filter_professional = st.selectbox(
        "Profesional",
        ["Todos"] + [p['name'] for p in professionals]
    )
    
    filter_status = st.selectbox(
        "Estado",
        ["Todos", "Confirmada", "Pendiente", "Cancelada"]
    )
    
    filter_payment = st.selectbox(
        "Pago",
        ["Todos", "Pagado", "Anticipo pagado", "Pendiente"]
    )
    
    st.markdown("---")
    
    st.markdown("### ⚡ Acciones")
    st.markdown("---")
    
    # BOTÓN 1: Actualizar
    if st.button("🔄 Actualizar", use_container_width=True, key="btn_refresh"):
        st.cache_data.clear()
        st.rerun()
    
    # BOTÓN 2: Exportar a Excel
    if st.button("📥 Exportar a Excel", use_container_width=True, key="btn_export"):
        st.session_state.current_action = 'export_excel'
        st.rerun()
    
    # BOTÓN 3: Enviar Recordatorios
    if st.button("📧 Enviar Recordatorios", use_container_width=True, key="btn_reminders"):
        st.session_state.current_action = 'send_reminders'
        st.rerun()
    
    # BOTÓN 4: Cerrar Sesión
    if st.button("🚪 Cerrar Sesión", use_container_width=True, key="btn_logout"):
        st.session_state.clear()
        st.info("✅ Sesión cerrada. Recargando...")
        st.rerun()
    
    # ========== PROCESAR ACCIONES ==========
    
    # Exportar a Excel
    if st.session_state.get('current_action') == 'export_excel':
        with st.spinner("⏳ Generando Excel..."):
            excel_buffer = export_bookings_to_excel()
            if excel_buffer:
                st.download_button(
                    label="💾 Descargar Excel",
                    data=excel_buffer,
                    file_name=f"citas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="download_excel"
                )
                st.success("✅ Excel generado exitosamente")
                st.session_state.current_action = None
            else:
                st.error("❌ Error generando Excel")
                st.session_state.current_action = None
    
    # Enviar Recordatorios
    if st.session_state.get('current_action') == 'send_reminders':
        with st.spinner("⏳ Enviando recordatorios..."):
            sent_count = send_appointment_reminders()
            if sent_count > 0:
                st.success(f"✅ {sent_count} recordatorios enviados correctamente")
            elif sent_count == 0:
                st.info("ℹ️ No hay citas para recordar mañana")
            else:
                st.error("❌ Error enviando recordatorios")
            st.session_state.current_action = None


# ==================== VISTA PRINCIPAL ====================

# ===== SECCIÓN CORREGIDA: if view_mode == "📅 Calendario del Día": =====
# REEMPLAZAR TODO desde: if view_mode == "📅 Calendario del Día":
# HASTA antes de: elif view_mode == "📊 Agenda Semanal":

if view_mode == "📅 Calendario del Día":
    st.markdown("## 📅 Calendario del Día - Dashboard Mejorado")
    
    selected_date_str = st.session_state.selected_date.strftime('%Y-%m-%d')
    
    # Obtener todas las reservas del día
    with db.get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT b.*, p.name as professional_name
            FROM bookings b
            LEFT JOIN professionals p ON b.professional_id = p.id
            WHERE b.date = %s
            ORDER BY b.start_time
        """, (selected_date_str,))
        bookings = [dict(zip([desc[0] for desc in cursor.description], row)) for row in cursor.fetchall()]
    
    # Calcular estadísticas
    stats = calculate_stats(bookings)
    
    # Mostrar métricas principales
    col1, col2, col3, col4, col5, col6 = st.columns(6)
    
    with col1:
        st.metric("Total Citas", stats['total_bookings'])
    
    with col2:
        st.metric("Confirmadas", stats['confirmed'])
    
    with col3:
        st.metric("Pendientes", stats['pending'])
    
    with col4:
        st.metric("Ingresos", f"${stats['total_revenue']:,.0f}")
    
    with col5:
        st.metric("Anticipos", f"${stats['deposits_collected']:,.0f}")
    
    with col6:
        st.metric("Pendiente", f"${stats['pending_payments']:,.0f}")
    
    st.markdown("---")
    
    # Agrupar por profesional
    professionals_data = {}
    for booking in bookings:
        prof_name = booking['professional_name'] or 'Sin asignar'
        if prof_name not in professionals_data:
            professionals_data[prof_name] = []
        professionals_data[prof_name].append(booking)
    
    # Aplicar filtros
    if filter_professional != "Todos":
        professionals_data = {k: v for k, v in professionals_data.items() if k == filter_professional}
    
    if filter_status != "Todos":
        status_map = {'Confirmada': 'confirmed', 'Pendiente': 'pending', 'Cancelada': 'cancelled'}
        for prof in professionals_data:
            professionals_data[prof] = [b for b in professionals_data[prof] if b['status'] == status_map[filter_status]]
    
    if filter_payment != "Todos":
        payment_map = {'Pagado': 'paid', 'Anticipo pagado': 'partial', 'Pendiente': 'pending'}
        for prof in professionals_data:
            filtered_bookings = []
            for b in professionals_data[prof]:
                _, payment_status = get_payment_status(b['total_price'], b['deposit_paid'])
                status_key = 'paid' if b['deposit_paid'] >= b['total_price'] else ('partial' if b['deposit_paid'] > 0 else 'pending')
                if status_key == payment_map[filter_payment]:
                    filtered_bookings.append(b)
            professionals_data[prof] = filtered_bookings
    
    # === NUEVA SECCIÓN: ANÁLISIS DE OCUPACIÓN ===
    st.markdown("### 📊 Análisis de Ocupación")

    # Inicializar variable ✅ IMPORTANTE
    df_occupation = None

    # Calcular ocupación por profesional
    occupation_data = []
    for prof_name, prof_bookings in professionals_data.items():
        if not prof_bookings:
            continue
        
        total_bookings = len(prof_bookings)
        confirmed_bookings = len([b for b in prof_bookings if b['status'] == 'confirmed'])
        
        # Conversión segura de precios
        try:
            prices = []
            for b in prof_bookings:
                price = b.get('total_price', 0)
                prices.append(convert_to_native(price))
            total_revenue = sum(float(p) if p else 0 for p in prices)
        except Exception as e:
            st.error(f"❌ Error procesando precios: {e}")
            total_revenue = 0
        
        if total_bookings > 0:
            occupation_rate = (confirmed_bookings / total_bookings) * 100
        else:
            occupation_rate = 0
        
        occupation_data.append({
            'Profesional': convert_to_native(prof_name),
            'Citas': convert_to_native(total_bookings),
            'Confirmadas': convert_to_native(confirmed_bookings),
            'Ocupación %': convert_to_native(occupation_rate),
            'Ingresos': convert_to_native(total_revenue)
        })

    if occupation_data:
        try:
            # Convertir lista completa de datos
            occupation_data_clean = [convert_to_native(row) for row in occupation_data]
            
            # Crear DataFrame
            df_occupation = pd.DataFrame(occupation_data_clean)
            
            # Gráfico de ocupación por profesional
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("#### 👥 Citas por Profesional")
                chart_data = df_occupation.set_index('Profesional')['Citas'].to_dict()
                st.bar_chart(chart_data)
            
            with col2:
                st.markdown("#### 📈 Tasa de Ocupación (%)")
                chart_data = df_occupation.set_index('Profesional')['Ocupación %'].to_dict()
                st.bar_chart(chart_data)
            
            st.markdown("---")
            
            # Tabla de resumen de ocupación
            st.markdown("#### 📋 Resumen de Ocupación")
            st.dataframe(
                df_occupation.sort_values('Ingresos', ascending=False),
                width='stretch',
                hide_index=True
            )
        
        except Exception as e:
            st.error(f"❌ Error creando gráficos de ocupación: {e}")
            st.write("**Detalles del error:**")
            st.write(str(e))
            st.write("**Datos problemáticos:**")
            for i, row in enumerate(occupation_data):
                st.write(f"Fila {i}: {row}")
            df_occupation = None

    else:
        st.info("📊 No hay datos de ocupación para mostrar")

    
    # === NUEVA SECCIÓN: HORAS PICO ===
    st.markdown("### ⏰ Análisis de Horas")
    
    # Agrupar citas por hora
    hours_data = {}
    for booking in bookings:
        hour = str(booking['start_time']).split(':')[0]
        if hour not in hours_data:
            hours_data[hour] = 0
        hours_data[hour] += 1
    
    if hours_data:
        hours_df = pd.DataFrame(list(hours_data.items()), columns=['Hora', 'Citas'])
        hours_df = hours_df.sort_values('Hora')
        
        st.markdown("#### 🕐 Distribución de Citas por Hora")
        st.bar_chart(hours_df.set_index('Hora'))
    else:
        st.info("No hay citas para este día")
    
    st.markdown("---")
    
    # === DETALLE POR PROFESIONAL ===
    st.markdown("### 👤 Detalle por Profesional")
    
    for professional, bookings_list in professionals_data.items():
        with st.expander(f"👤 {professional} ({len(bookings_list)} citas)"):
            if bookings_list:
                # Timeline view
                st.markdown("#### ⏰ Horario del Día")
                
                for booking in sorted(bookings_list, key=lambda x: x['start_time']):
                    col1, col2, col3, col4, col5 = st.columns([1, 2, 2, 2, 2])
                    
                    with col1:
                        st.markdown(f"**{format_time_range(booking['start_time'], booking['end_time'])}**")
                    
                    with col2:
                        st.markdown(f"👤 {booking['client_name']}")
                    
                    with col3:
                        st.markdown(get_status_badge(booking['status']), unsafe_allow_html=True)
                    
                    with col4:
                        # Calcular y mostrar estado de pago
                        badge_html = get_payment_badge_from_amounts(booking['total_price'], booking['deposit_paid'])
                        st.markdown(badge_html, unsafe_allow_html=True)
                    
                    with col5:
                        if st.button("👁️", key=f"view_{booking['id']}"):
                            st.write(f"Detalles de {booking['booking_code']}")
                
                st.markdown("---")
                
                # Resumen del profesional
                prof_revenue = sum(b['total_price'] for b in bookings_list)
                prof_deposits = sum(b['deposit_paid'] for b in bookings_list)
                
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Total Citas", len(bookings_list))
                with col2:
                    st.metric("Ingresos", f"${prof_revenue:,.0f}")
                with col3:
                    st.metric("Anticipos", f"${prof_deposits:,.0f}")
                with col4:
                    st.metric("Pendiente", f"${prof_revenue - prof_deposits:,.0f}")
            else:
                st.info("Sin citas para este día")

# ===== REEMPLAZAR SECCIÓN: elif view_mode == "📊 Agenda Semanal": =====
# Copiar y reemplazar desde esta línea hasta antes de elif view_mode == "👥 CRM Clientes":

elif view_mode == "📊 Agenda Semanal":
    st.markdown("## 📊 Vista Semanal")
    
    # Calcular rango de la semana
    start_of_week = selected_date - timedelta(days=selected_date.weekday())
    week_dates = [start_of_week + timedelta(days=i) for i in range(7)]
    
    # Crear columnas para cada día
    day_cols = st.columns(7)
    
    weekdays_es = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
    
    for idx, date in enumerate(week_dates):
        with day_cols[idx]:
            date_str = date.strftime('%Y-%m-%d')
            bookings = db.get_daily_bookings(date_str)
            
            is_today = date == datetime.now().date()
            bg_color = "#DBEAFE" if is_today else "#F3F4F6"
            
            st.markdown(f"""
            <div style='background: {bg_color}; padding: 1rem; border-radius: 10px; text-align: center; margin-bottom: 1rem;'>
                <div style='font-weight: bold; color: #1F2937;'>{weekdays_es[idx]}</div>
                <div style='font-size: 1.5rem; font-weight: bold; color: #3B82F6;'>{date.day}</div>
                <div style='font-size: 0.8rem; color: #6B7280;'>{len(bookings)} citas</div>
            </div>
            """, unsafe_allow_html=True)
            
            if bookings:
                # Mostrar TODAS las citas del día (no solo 3)
                for booking in sorted(bookings, key=lambda x: x['start_time']):
                    # Obtener servicios de la cita
                    with db.get_connection() as conn:
                        cursor = conn.cursor()
                        cursor.execute('''
                            SELECT s.name, s.duration
                            FROM booking_services bs
                            JOIN services s ON bs.service_id = s.id
                            WHERE bs.booking_id = %s
                            ORDER BY s.name
                        ''', (booking['id'],))
                        services = [dict(zip([desc[0] for desc in cursor.description], row)) for row in cursor.fetchall()]
                    
                    # Crear string de servicios
                    services_str = ", ".join([s['name'] for s in services]) if services else "Sin servicios"
                    
                    # Determinar color según estado
                    status_colors = {
                        'confirmed': '#D1FAE5',  # Verde
                        'pending': '#FEF3C7',    # Amarillo
                        'cancelled': '#FEE2E2'   # Rojo
                    }
                    status_color = status_colors.get(booking['status'], '#F3F4F6')
                    
                    st.markdown(f"""
                    <div style='background: {status_color}; padding: 0.75rem; border-radius: 8px; margin-bottom: 0.5rem; font-size: 0.75rem; border-left: 3px solid #3B82F6;'>
                        <div style='font-weight: bold; margin-bottom: 0.25rem;'>{booking['start_time']}</div>
                        <div style='margin-bottom: 0.25rem;'><strong>👤</strong> {booking['client_name']}</div>
                        <div style='margin-bottom: 0.25rem;'><strong>💅</strong> {services_str}</div>
                        <div style='color: #666;'><strong>👥</strong> {booking['professional_name']}</div>
                    </div>
                    """, unsafe_allow_html=True)
            else:
                st.info("Sin citas")
    
    st.markdown("---")
    
    # Resumen de la semana
    st.markdown("### 📊 Resumen Semanal")
    
    week_stats = {
        'total_bookings': 0,
        'confirmed': 0,
        'pending': 0,
        'total_revenue': 0,
        'deposits_collected': 0
    }
    
    for date in week_dates:
        date_str = date.strftime('%Y-%m-%d')
        bookings = db.get_daily_bookings(date_str)
        week_stats['total_bookings'] += len(bookings)
        week_stats['confirmed'] += len([b for b in bookings if b['status'] == 'confirmed'])
        week_stats['pending'] += len([b for b in bookings if b['status'] == 'pending'])
        week_stats['total_revenue'] += sum(b['total_price'] for b in bookings)
        week_stats['deposits_collected'] += sum(b['deposit_paid'] for b in bookings)
    
    col1, col2, col3, col4, col5 = st.columns(5)
    
    with col1:
        st.metric("Total Citas", week_stats['total_bookings'])
    
    with col2:
        st.metric("Confirmadas", week_stats['confirmed'])
    
    with col3:
        st.metric("Pendientes", week_stats['pending'])
    
    with col4:
        st.metric("Ingresos", f"${week_stats['total_revenue']:,.0f}")
    
    with col5:
        st.metric("Anticipos", f"${week_stats['deposits_collected']:,.0f}")

# ===== REEMPLAZAR SECCIÓN: elif view_mode == "👥 CRM Clientes": =====
# Copiar y reemplazar desde esta línea hasta antes de elif view_mode == "📈 Reportes":

elif view_mode == "💳 Gestión de Pagos":
    st.markdown("## 💳 Gestión de Pagos")
    
    # Filtros
    col1, col2, col3 = st.columns(3)
    
    with col1:
        payment_status_filter = st.selectbox(
            "Estado de Pago",
            ["Todos", "Pendiente", "Anticipo Pagado", "Pagado Completo"]
        )
    
    with col2:
        date_range = st.selectbox(
            "Período",
            ["Hoy", "Esta Semana", "Este Mes", "Últimos 30 días", "Personalizado"]
        )
    
    with col3:
        min_amount = st.number_input("Monto Mínimo ($)", min_value=0.0, value=0.0)
    
    # Obtener pagos según filtros
    with db.get_connection() as conn:
        cursor = conn.cursor()
        
        # Construir query dinámico
        query = '''
            SELECT 
                b.id,
                b.booking_code,
                b.client_name,
                b.client_phone,
                b.client_email,
                b.date,
                b.start_time,
                b.total_price,
                b.deposit_paid,
                b.status as booking_status,
                p.mercado_pago_id,
                p.payment_status,
                p.verified,
                p.created_at as payment_date,
                pr.name as professional_name
            FROM bookings b
            LEFT JOIN payments p ON b.id = p.booking_id
            LEFT JOIN professionals pr ON b.professional_id = pr.id
            WHERE 1=1
        '''
        
        params = []
        
        # Filtro por estado de pago
        if payment_status_filter == "Pendiente":
            query += " AND b.status = 'pending'"
        elif payment_status_filter == "Anticipo Pagado":
            query += " AND b.status = 'confirmed'"
        elif payment_status_filter == "Pagado Completo":
            query += " AND b.deposit_paid >= b.total_price"
        
        # Filtro por rango de fechas
        if date_range == "Hoy":
            query += " AND b.date = CURRENT_DATE"
        elif date_range == "Esta Semana":
            query += " AND b.date >= CURRENT_DATE - INTERVAL '7 days'"
        elif date_range == "Este Mes":
            query += " AND DATE_TRUNC('month', b.date) = DATE_TRUNC('month', CURRENT_DATE)"
        elif date_range == "Últimos 30 días":
            query += " AND b.date >= CURRENT_DATE - INTERVAL '30 days'"
        
        # Filtro por monto mínimo
        if min_amount > 0:
            query += " AND (b.total_price - b.deposit_paid) >= %s"
            params.append(min_amount)
        
        query += " ORDER BY b.date DESC, b.start_time DESC"
        
        cursor.execute(query, params)
        payments = [dict(zip([desc[0] for desc in cursor.description], row)) for row in cursor.fetchall()]
    
    # Calcular estadísticas
    total_pending = sum(max(0, p['total_price'] - p['deposit_paid']) for p in payments)
    total_collected = sum(p['deposit_paid'] for p in payments)
    total_citas = len(payments)
    
    # Métricas
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("Total Citas", total_citas)
    
    with col2:
        st.metric("Pendiente de Cobro", f"${total_pending:,.2f}")
    
    with col3:
        st.metric("Cobrado", f"${total_collected:,.2f}")
    
    with col4:
        total_value = sum(p['total_price'] for p in payments)
        st.metric("Valor Total", f"${total_value:,.2f}")
    
    st.markdown("---")
    
    # Tabla de pagos
    if payments:
        st.markdown("### 📋 Detalle de Pagos")
        
        # Crear lista para mostrar
        payment_list = []
        for p in payments:
            pending = max(0, p['total_price'] - p['deposit_paid'])
            payment_status_text = "✅ Pagado" if pending <= 0 else "⏳ Pendiente"
            
            payment_list.append({
                'Cita': p['booking_code'],
                'Cliente': p['client_name'],
                'Profesional': p['professional_name'] or 'N/A',
                'Fecha': p['date'],
                'Hora': p['start_time'],
                'Total': f"${p['total_price']:,.2f}",
                'Pagado': f"${p['deposit_paid']:,.2f}",
                'Pendiente': f"${pending:,.2f}",
                'Estado': payment_status_text,
                'ID': p['id']
            })
        
        # Mostrar en expandibles
        for payment in payment_list:
            col1, col2, col3, col4, col5 = st.columns([2, 2, 2, 2, 1])
            
            with col1:
                st.markdown(f"**{payment['Cita']}**")
                st.caption(f"{payment['Cliente']}")
            
            with col2:
                st.markdown(f"📅 {payment['Fecha']}")
                st.caption(f"⏰ {payment['Hora']}")
            
            with col3:
                st.markdown(f"Total: {payment['Total']}")
                st.caption(f"Pagado: {payment['Pagado']}")
            
            with col4:
                if float(payment['Pendiente'].replace('$', '').replace(',', '')) > 0:
                    st.warning(f"Pendiente: {payment['Pendiente']}")
                else:
                    st.success(f"Completo ✅")
                st.caption(f"👥 {payment['Profesional']}")
            
            with col5:
                if st.button("✏️", key=f"edit_{payment['ID']}", help="Validar/editar pago"):
                    st.session_state.selected_payment_id = payment['ID']
                    st.session_state.show_payment_form = True
        
        st.markdown("---")
        
        # Formulario para validar pago manualmente
        if 'show_payment_form' in st.session_state and st.session_state.show_payment_form:
            st.markdown("### ✏️ Validar Pago Manualmente")
            
            selected_payment = next((p for p in payments if p['id'] == st.session_state.selected_payment_id), None)
            
            if selected_payment:
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown(f"**Cita:** {selected_payment['booking_code']}")
                    st.markdown(f"**Cliente:** {selected_payment['client_name']}")
                    st.markdown(f"**Total:** ${selected_payment['total_price']:,.2f}")
                
                with col2:
                    st.markdown(f"**Fecha:** {selected_payment['date']}")
                    st.markdown(f"**Hora:** {selected_payment['start_time']}")
                    st.markdown(f"**Pagado:** ${selected_payment['deposit_paid']:,.2f}")
                
                st.markdown("---")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    operation_number = st.text_input(
                        "Número de Operación (Mercado Pago)",
                        placeholder="Ej: 12345678901",
                        key=f"operation_{selected_payment['id']}"
                    )
                
                with col2:
                    amount_paid = st.number_input(
                        "Monto Pagado ($)",
                        min_value=0.0,
                        value=selected_payment['total_price'] - selected_payment['deposit_paid'],
                        key=f"amount_{selected_payment['id']}"
                    )
                
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    if st.button("✅ Validar Pago", width='stretch', key=f"validate_{selected_payment['id']}"):
                        if operation_number:
                            # Obtener access token
                            try:
                                access_token = st.secrets["mercadopago"]["access_token"]
                            except (KeyError, TypeError):
                                st.error("Token de Mercado Pago no configurado")
                                access_token = None
                            
                            if access_token:
                                # Validar pago
                                is_valid, payment_data, error = db.validate_mercadopago_payment(
                                    operation_number,
                                    selected_payment['id'],
                                    access_token
                                )
                                
                                if is_valid:
                                    # Confirmar pago
                                    success, msg = db.confirm_payment_with_operation(
                                        selected_payment['booking_code'],
                                        operation_number,
                                        payment_data
                                    )
                                    
                                    if success:
                                        st.success(f"✅ {msg}")
                                        st.session_state.show_payment_form = False
                                        st.rerun()
                                    else:
                                        st.error(f"❌ Error: {msg}")
                                else:
                                    st.error(f"❌ {error}")
                        else:
                            st.error("Por favor ingresa el número de operación")
                
                with col2:
                    if st.button("📝 Registrar Manual", width='stretch', key=f"manual_{selected_payment['id']}"):
                        # Registrar pago manual sin validar con MP
                        if amount_paid > 0:
                            st.info(f"Pago de ${amount_paid:,.2f} registrado manualmente")
                            st.session_state.show_payment_form = False
                            st.rerun()
                        else:
                            st.error("El monto debe ser mayor a 0")
                
                with col3:
                    if st.button("❌ Cancelar", width='stretch', key=f"cancel_{selected_payment['id']}"):
                        st.session_state.show_payment_form = False
                        st.rerun()
    
    else:
        st.info("No hay pagos que coincidan con los filtros seleccionados")

elif view_mode == "📈 Reportes":
    st.markdown("## 📈 Reportes y Estadísticas")
    
    # Selector de rango de fechas
    col1, col2 = st.columns(2)
    with col1:
        start_date = st.date_input("Fecha inicial", datetime.now() - timedelta(days=30))
    with col2:
        end_date = st.date_input("Fecha final", datetime.now())
    
    st.markdown("---")
    
    # Obtener todas las reservas en el rango
    with db.get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                b.date::date as booking_date,
                COUNT(*) as total_bookings,
                SUM(b.total_price) as total_revenue,
                SUM(b.deposit_paid) as deposits_collected
            FROM bookings b
            WHERE b.date BETWEEN %s AND %s
            GROUP BY b.date::date
            ORDER BY b.date
        """, (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')))
        
        #results = cursor.fetchall()
        # Convertir cada fila a diccionario y asegurar tipos
        results = []
        for row in cursor.fetchall():
            row_dict = dict(zip([desc[0] for desc in cursor.description], row))
            # Convertir Decimal a float si es necesario
            for key in row_dict:
                if isinstance(row_dict[key], (Decimal, float)):
                    row_dict[key] = float(row_dict[key])
            results.append(row_dict)
    
    if results:
        # Convertir a DataFrame
        df = pd.DataFrame(results)
        
        # Renombrar columnas
        df.rename(columns={
            'booking_date': 'Fecha',
            'total_bookings': 'Citas',
            'total_revenue': 'Ingresos',
            'deposits_collected': 'Anticipos'
        }, inplace=True)
        
        # Asegurar tipos de datos
        df['Fecha'] = pd.to_datetime(df['Fecha']).dt.strftime('%Y-%m-%d')
        df['Citas'] = df['Citas'].astype(int)
        df['Ingresos'] = df['Ingresos'].astype(float)
        df['Anticipos'] = df['Anticipos'].astype(float)
        
        # Métricas generales
        col1, col2, col3 = st.columns(3)
        
        with col1:
            total_bookings = df['Citas'].sum()
            st.metric("Total de Citas", f"{int(total_bookings)}")
        
        with col2:
            total_revenue = df['Ingresos'].sum()
            st.metric("Ingresos Totales", f"${total_revenue:,.2f}")
        
        with col3:
            avg_ticket = total_revenue / total_bookings if total_bookings > 0 else 0
            st.metric("Ticket Promedio", f"${avg_ticket:,.2f}")
        
        st.markdown("---")
        
        # Gráfico de ingresos
        st.markdown("### 💰 Ingresos por Día")
        st.line_chart(df.set_index('Fecha')['Ingresos'])
        
        st.markdown("### 📊 Citas por Día")
        st.bar_chart(df.set_index('Fecha')['Citas'])
        
        # Tabla detallada
        st.markdown("### 📋 Detalle por Fecha")
        st.dataframe(df, use_container_width=True)
    else:
        st.info("No hay datos para el rango seleccionado")

# ===== REEMPLAZAR SECCIÓN: elif view_mode == "⚙️ Configuración": =====
# ===== REEMPLAZAR SECCIÓN COMPLETA: elif view_mode == "⚙️ Configuración": =====
# Copiar TODO desde aquí hasta el final del archivo

elif view_mode == "⚙️ Configuración":
    st.markdown("## ⚙️ Centro de Control - Configuración del Sistema")
    
    tab_list = [
        "🛡️ Gestión de Usuarios", # ⬅️ NUEVA PESTAÑA
        "👥 Profesionales", 
        "💅 Servicios", 
        "🔗 Profesional-Servicio",
        "⏰ Horarios",
        "📋 Respaldo"
    ]
    tab_objs = st.tabs(tab_list)
    
    # Asignar a variables para fácil acceso
    tab0 = tab_objs[0]
    tab1 = tab_objs[1]
    tab2 = tab_objs[2]
    tab3 = tab_objs[3]
    tab4 = tab_objs[4]
    tab5 = tab_objs[5]
    
    # ===== TAB 0: GESTIÓN DE USUARIOS =====
    with tab0:
        st.markdown("### 🛡️ Gestión de Usuarios")
        
        # --- Formulario de nuevo usuario ---
        with st.expander("➕ Crear Nuevo Usuario"):
            new_username = st.text_input("Username (Ej: admin)", key="new_user_username")
            new_user_name = st.text_input("Nombre Completo", key="new_user_name")
            new_password = st.text_input("Contraseña", type="password", key="new_user_password")
            
            if st.button("✅ Guardar Usuario", width='stretch'):
                if new_username and new_user_name and new_password:
                    success, message = db.create_user(new_username, new_password, new_user_name)
                    if success:
                        st.success(message)
                        st.rerun()
                    else:
                        st.error(message)
                else:
                    st.error("❌ Todos los campos son requeridos")
        
        st.markdown("---")
        
        # --- Listar y gestionar usuarios ---
        st.markdown("#### 📋 Usuarios del Sistema")
        
        users = db.get_all_users()
        
        if users:
            for user in users:
                col1, col2 = st.columns([4, 1])
                
                with col1:
                    st.markdown(f"**{user['name']}**")
                    st.caption(f"👤 Username: `{user['username']}` | ID: `{user['id']}`")
                
                with col2:
                    if st.button("🔑", key=f"reset_pass_{user['id']}", help="Cambiar Contraseña"):
                        st.session_state.show_reset_pass_form = user['id']
                
                # Formulario para cambiar contraseña
                if st.session_state.get('show_reset_pass_form') == user['id']:
                    st.markdown("##### 🔑 Cambiar Contraseña")
                    reset_pass = st.text_input("Nueva Contraseña", type="password", key=f"new_pass_{user['id']}")
                    
                    col1_r, col2_r = st.columns(2)
                    
                    with col1_r:
                        if st.button("💾 Guardar Contraseña", width='stretch', key=f"save_pass_{user['id']}"):
                            if reset_pass:
                                success, message = db.update_password(user['username'], reset_pass)
                                if success:
                                    st.success(message)
                                    st.session_state.show_reset_pass_form = None
                                    st.rerun()
                                else:
                                    st.error(message)
                            else:
                                st.error("❌ La contraseña no puede estar vacía")
                    
                    with col2_r:
                        # Botón de eliminación
                        if st.button("🗑️ Eliminar Usuario", width='stretch', key=f"delete_user_{user['id']}"):
                            st.session_state.confirm_delete_user = user['id']

                        if st.session_state.get('confirm_delete_user') == user['id']:
                            st.warning(f"⚠️ ¿Eliminar a {user['name']}? Esta acción no se puede deshacer.")
                            col1_d, col2_d = st.columns(2)
                            with col1_d:
                                if st.button("✅ Sí, eliminar", key=f"confirm_del_user_{user['id']}"):
                                    success, message = db.delete_user(user['id'])
                                    if success:
                                        st.success(message)
                                        st.session_state.confirm_delete_user = None
                                        st.session_state.show_reset_pass_form = None
                                        st.rerun()
                                    else:
                                        st.error(message)
                            with col2_d:
                                if st.button("❌ Cancelar", key=f"cancel_del_user_{user['id']}"):
                                    st.session_state.confirm_delete_user = None
                
                st.markdown("---")
        else:
            st.info("No hay usuarios registrados. Crea el primer usuario administrador.")
    
    # ===== TAB 1: PROFESIONALES =====
    with tab1:
        st.markdown("### 👥 Gestión de Profesionales")
        
        col1, col2 = st.columns([3, 1])
        with col2:
            if st.button("➕ Nuevo Profesional", width='stretch'):
                st.session_state.show_new_professional_form = True
        
        # Formulario para nuevo profesional
        if st.session_state.get('show_new_professional_form', False):
            st.markdown("#### ➕ Agregar Nuevo Profesional")
            
            col1, col2 = st.columns(2)
            with col1:
                new_name = st.text_input("Nombre Completo", key="new_prof_name")
                new_specialization = st.text_input("Especialización", key="new_prof_specialization")
                new_phone = st.text_input("Teléfono", key="new_prof_phone")
            
            with col2:
                new_email = st.text_input("Email", key="new_prof_email")
                new_active = st.checkbox("Activo", value=True, key="new_prof_active")
            
            col1, col2, col3 = st.columns([1, 1, 2])
            
            with col1:
                if st.button("✅ Guardar", width='stretch', key="save_new_prof"):
                    if new_name and new_specialization:
                        with db.get_connection() as conn:
                            cursor = conn.cursor()
                            cursor.execute('''
                                INSERT INTO professionals (name, specialization, phone, email, active)
                                VALUES (%s, %s, %s, %s, %s)
                            ''', (new_name, new_specialization, new_phone, new_email, new_active))
                            conn.commit()
                        st.success("✅ Profesional creado exitosamente")
                        st.session_state.show_new_professional_form = False
                        st.rerun()
                    else:
                        st.error("❌ Nombre y Especialización son requeridos")
            
            with col2:
                if st.button("❌ Cancelar", width='stretch', key="cancel_new_prof"):
                    st.session_state.show_new_professional_form = False
                    st.rerun()
        
        st.markdown("---")
        
        # Listar profesionales
        st.markdown("#### 📋 Profesionales Registrados")
        
        with db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM professionals ORDER BY name")
            professionals = [dict(zip([desc[0] for desc in cursor.description], row)) for row in cursor.fetchall()]
        
        if professionals:
            for prof in professionals:
                col1, col2 = st.columns([4, 1])
                
                with col1:
                    st.markdown(f"**{prof['name']}**")
                    st.caption(f"🎓 {prof['specialization']} | 📱 {prof['phone']} | 📧 {prof['email']}")
                    status = "✅ Activo" if prof['active'] else "⭕ Inactivo"
                    st.caption(f"Estado: {status}")
                
                with col2:
                    if st.button("✏️", key=f"edit_prof_{prof['id']}", help="Editar"):
                        st.session_state.selected_prof_id = prof['id']
                        st.session_state.show_edit_prof_form = True
                
                # Formulario de edición
                if st.session_state.get('show_edit_prof_form') and st.session_state.get('selected_prof_id') == prof['id']:
                    st.markdown("##### ✏️ Editar Profesional")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        edit_name = st.text_input("Nombre", prof['name'], key=f"edit_name_{prof['id']}")
                        edit_specialization = st.text_input("Especialización", prof['specialization'], key=f"edit_spec_{prof['id']}")
                        edit_phone = st.text_input("Teléfono", prof['phone'], key=f"edit_phone_{prof['id']}")
                    
                    with col2:
                        edit_email = st.text_input("Email", prof['email'], key=f"edit_email_{prof['id']}")
                        edit_active = st.checkbox("Activo", value=prof['active'], key=f"edit_active_{prof['id']}")
                    
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        if st.button("💾 Guardar", width='stretch', key=f"save_edit_{prof['id']}"):
                            with db.get_connection() as conn:
                                cursor = conn.cursor()
                                cursor.execute('''
                                    UPDATE professionals 
                                    SET name=%s, specialization=%s, phone=%s, email=%s, active=%s
                                    WHERE id=%s
                                ''', (edit_name, edit_specialization, edit_phone, edit_email, edit_active, prof['id']))
                                conn.commit()
                            st.success("✅ Cambios guardados")
                            st.session_state.show_edit_prof_form = False
                            st.rerun()
                    
                    with col2:
                        if st.button("❌ Cancelar", width='stretch', key=f"cancel_edit_{prof['id']}"):
                            st.session_state.show_edit_prof_form = False
                            st.rerun()
                    
                    with col3:
                        if st.button("🗑️ Eliminar", width='stretch', key=f"delete_prof_{prof['id']}"):
                            st.session_state.confirm_delete_prof = prof['id']
                    
                    # Confirmación de eliminación
                    if st.session_state.get('confirm_delete_prof') == prof['id']:
                        st.warning(f"⚠️ ¿Eliminar a {prof['name']}? Esta acción no se puede deshacer.")
                        col1, col2 = st.columns(2)
                        with col1:
                            if st.button("✅ Sí, eliminar", key=f"confirm_delete_{prof['id']}"):
                                with db.get_connection() as conn:
                                    cursor = conn.cursor()
                                    cursor.execute("DELETE FROM professionals WHERE id=%s", (prof['id'],))
                                    conn.commit()
                                st.success("✅ Profesional eliminado")
                                st.session_state.confirm_delete_prof = None
                                st.session_state.show_edit_prof_form = False
                                st.rerun()
                        with col2:
                            if st.button("❌ Cancelar", key=f"cancel_delete_{prof['id']}"):
                                st.session_state.confirm_delete_prof = None
                
                st.markdown("---")
        else:
            st.info("No hay profesionales registrados. Crea uno para comenzar.")
    
    # ===== TAB 2: SERVICIOS =====
    # ============================================
# REEMPLAZAR en admin.py: Sección de Gestión de Servicios
# ============================================

# ===== TAB 2: SERVICIOS (ACTUALIZADA) =====
    with tab2:
        st.markdown("### 🧴 Gestión de Servicios")
        
        service_mode = st.radio(
            "Elige una opción",
            ["➕ Agregar Servicio", "✏️ Editar Servicio", "📋 Ver Servicios", "🔧 Gestionar Categorías"],
            horizontal=True,
            key="service_mode"
        )
        
        # ===== MODO 1: AGREGAR SERVICIO =====
        if service_mode == "➕ Agregar Servicio":
            st.markdown("#### ➕ Nuevo Servicio")
            
            # PASO 1: Seleccionar o crear categoría
            st.markdown("**Paso 1: Categoría**")
            
            existing_categories = db.get_active_categories()
            category_options = {cat['name']: cat['id'] for cat in existing_categories}
            
            col1, col2 = st.columns([3, 1])
            
            with col1:
                selected_category_name = st.selectbox(
                    "Selecciona una categoría existente",
                    options=list(category_options.keys()),
                    key="select_category"
                )
                selected_category_id = category_options[selected_category_name]
            
            with col2:
                if st.button("➕ Nueva Categoría", key="new_category_btn"):
                    st.session_state.show_new_category = True
            
            # Formulario para nueva categoría (si se activa)
            if st.session_state.get('show_new_category', False):
                st.markdown("---")
                st.markdown("**Crear Nueva Categoría**")
                
                new_cat_col1, new_cat_col2, new_cat_col3 = st.columns(3)
                
                with new_cat_col1:
                    new_category_name = st.text_input(
                        "Nombre de la categoría",
                        placeholder="ej: Uñas",
                        key="new_cat_name"
                    )
                
                with new_cat_col2:
                    new_category_icon = st.selectbox(
                        "Ícono",
                        ["💅", "💇", "💆", "✨", "🎨", "🪮", "📋", "💄"],
                        key="new_cat_icon"
                    )
                
                with new_cat_col3:
                    new_category_color = st.color_picker(
                        "Color",
                        value="#EC4899",
                        key="new_cat_color"
                    )
                
                new_category_desc = st.text_area(
                    "Descripción (opcional)",
                    placeholder="Describe los servicios en esta categoría",
                    key="new_cat_desc"
                )
                
                col_create, col_cancel = st.columns(2)
                
                with col_create:
                    if st.button("✅ Crear Categoría", key="create_cat_btn"):
                        success, message, cat_id = db.create_category(
                            name=new_category_name,
                            description=new_category_desc,
                            icon=new_category_icon,
                            color=new_category_color
                        )
                        
                        if success:
                            st.success(message)
                            st.session_state.show_new_category = False
                            selected_category_id = cat_id
                            st.rerun()
                        else:
                            st.error(message)
                
                with col_cancel:
                    if st.button("❌ Cancelar", key="cancel_cat_btn"):
                        st.session_state.show_new_category = False
                        st.rerun()
            
            # PASO 2: Datos del servicio
            st.markdown("---")
            st.markdown("**Paso 2: Datos del Servicio**")
            
            service_col1, service_col2 = st.columns(2)
            
            with service_col1:
                service_name = st.text_input(
                    "Nombre del servicio",
                    placeholder="ej: Manicura completa",
                    key="service_name"
                )
                service_duration = st.number_input(
                    "Duración (minutos)",
                    min_value=15,
                    step=15,
                    value=60,
                    key="service_duration"
                )
                service_price = st.number_input(
                    "Precio (MXN)",
                    min_value=0.0,
                    step=50.0,
                    value=200.0,
                    key="service_price"
                )
            
            with service_col2:
                service_description = st.text_area(
                    "Descripción",
                    placeholder="Describe el servicio",
                    key="service_description"
                )
                service_deposit = st.number_input(
                    "Anticipo requerido (MXN)",
                    min_value=0.0,
                    step=50.0,
                    value=100.0,
                    key="service_deposit"
                )
            
            # PASO 3: Asignar profesionales
            st.markdown("---")
            st.markdown("**Paso 3: Profesionales (Opcional)**")
            
            with db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, name FROM professionals WHERE active = TRUE ORDER BY name")
                professionals = [dict(zip([desc[0] for desc in cursor.description], row)) for row in cursor.fetchall()]
            
            if professionals:
                selected_professionals = st.multiselect(
                    "Selecciona profesionales que pueden realizar este servicio",
                    options=[p['name'] for p in professionals],
                    key="service_professionals"
                )
                selected_prof_ids = [p['id'] for p in professionals if p['name'] in selected_professionals]
            else:
                st.info("ℹ️ No hay profesionales registrados aún")
                selected_prof_ids = []
            
            # BOTÓN CREAR
            st.markdown("---")
            if st.button("✅ Crear Servicio", use_container_width=True, type="primary", key="create_service"):
                if not service_name:
                    st.error("⚠️ El nombre del servicio es obligatorio")
                else:
                    try:
                        with db.get_connection() as conn:
                            cursor = conn.cursor()
                            
                            # Insertar servicio
                            cursor.execute("""
                                INSERT INTO services (name, description, price, duration, category_id, deposit, active)
                                VALUES (%s, %s, %s, %s, %s, %s, TRUE)
                                RETURNING id
                            """, (
                                service_name,
                                service_description,
                                service_price,
                                service_duration,
                                selected_category_id,
                                service_deposit
                            ))
                            
                            service_id = cursor.fetchone()[0]
                            
                            # Asignar profesionales si los hay
                            if selected_prof_ids:
                                for prof_id in selected_prof_ids:
                                    cursor.execute("""
                                        INSERT INTO professional_services (professional_id, service_id, active)
                                        VALUES (%s, %s, TRUE)
                                    """, (prof_id, service_id))
                            
                            conn.commit()
                        
                        st.success(f"✅ Servicio '{service_name}' creado exitosamente")
                        
                        # Limpiar formulario
                        st.session_state.show_new_category = False
                        st.rerun()
                    
                    except Exception as e:
                        st.error(f"❌ Error al crear servicio: {str(e)}")
        
        # ===== MODO 2: EDITAR SERVICIO =====
        elif service_mode == "✏️ Editar Servicio":
            st.markdown("#### ✏️ Editar Servicio")
            
            # Obtener servicios
            services = db.get_services()
            
            if services:
                service_options = {s['name']: s for s in services}
                selected_service_name = st.selectbox(
                    "Selecciona un servicio",
                    options=list(service_options.keys()),
                    key="edit_service_select"
                )
                
                selected_service = service_options[selected_service_name]
                
                st.markdown("---")
                
                # Obtener categoría actual
                categories = db.get_active_categories()
                category_options = {cat['name']: cat['id'] for cat in categories}
                
                edit_col1, edit_col2 = st.columns(2)
                
                with edit_col1:
                    edit_name = st.text_input(
                        "Nombre",
                        value=selected_service['name'],
                        key="edit_name"
                    )
                    edit_category = st.selectbox(
                        "Categoría",
                        options=list(category_options.keys()),
                        index=list(category_options.values()).index(selected_service.get('category_id')) 
                            if selected_service.get('category_id') in category_options.values() else 0,
                        key="edit_category"
                    )
                    edit_duration = st.number_input(
                        "Duración (minutos)",
                        value=selected_service['duration'],
                        min_value=15,
                        step=15,
                        key="edit_duration"
                    )
                
                with edit_col2:
                    edit_description = st.text_area(
                        "Descripción",
                        value=selected_service.get('description', ''),
                        key="edit_description"
                    )
                    edit_price = st.number_input(
                        "Precio (MXN)",
                        value=selected_service['price'],
                        min_value=0.0,
                        step=50.0,
                        key="edit_price"
                    )
                    edit_deposit = st.number_input(
                        "Anticipo (MXN)",
                        value=selected_service.get('deposit', 0),
                        min_value=0.0,
                        step=50.0,
                        key="edit_deposit"
                    )
                
                st.markdown("---")
                
                if st.button("✅ Actualizar Servicio", use_container_width=True, type="primary"):
                    try:
                        with db.get_connection() as conn:
                            cursor = conn.cursor()
                            cursor.execute("""
                                UPDATE services 
                                SET name = %s, description = %s, price = %s, duration = %s, 
                                    category_id = %s, deposit = %s
                                WHERE id = %s
                            """, (
                                edit_name,
                                edit_description,
                                edit_price,
                                edit_duration,
                                category_options[edit_category],
                                edit_deposit,
                                selected_service['id']
                            ))
                            conn.commit()
                        
                        st.success(f"✅ Servicio '{edit_name}' actualizado")
                        st.rerun()
                    
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")
            else:
                st.info("ℹ️ No hay servicios registrados")
        
        # ===== MODO 3: VER SERVICIOS =====
        elif service_mode == "📋 Ver Servicios":
            st.markdown("#### 📋 Lista de Servicios")
            
            # Filtrar por categoría
            categories = db.get_active_categories()
            
            if categories:
                selected_cat_filter = st.selectbox(
                    "Filtrar por categoría",
                    options=["Todas"] + [cat['name'] for cat in categories],
                    key="filter_category"
                )
                
                with db.get_connection() as conn:
                    cursor = conn.cursor()
                    
                    if selected_cat_filter == "Todas":
                        query = """
                            SELECT id, name, description, price, duration, category_id, deposit
                            FROM services
                            WHERE active = TRUE
                            ORDER BY category_id, name
                        """
                        cursor.execute(query)
                    else:
                        cat_id = [cat['id'] for cat in categories if cat['name'] == selected_cat_filter][0]
                        query = """
                            SELECT id, name, description, price, duration, category_id, deposit
                            FROM services
                            WHERE active = TRUE AND category_id = %s
                            ORDER BY name
                        """
                        cursor.execute(query, (cat_id,))
                    
                    services_list = [
                        dict(zip([desc[0] for desc in cursor.description], row))
                        for row in cursor.fetchall()
                    ]
                
                if services_list:
                    # Mostrar tabla
                    df = pd.DataFrame(services_list)
                    df = df.rename(columns={
                        'id': 'ID',
                        'name': 'Servicio',
                        'description': 'Descripción',
                        'price': 'Precio',
                        'duration': 'Duración (min)',
                        'category_id': 'Cat ID',
                        'deposit': 'Anticipo'
                    })
                    
                    st.dataframe(df, use_container_width=True)
                else:
                    st.info("ℹ️ No hay servicios en esta categoría")
        
        # ===== MODO 4: GESTIONAR CATEGORÍAS =====
        elif service_mode == "🔧 Gestionar Categorías":
            st.markdown("#### 🔧 Gestión de Categorías")
            
            cat_action = st.radio(
                "Elige una acción",
                ["➕ Agregar Categoría", "📊 Ver Categorías", "✏️ Editar Categoría", "🔍 Detectar Duplicadas"],
                horizontal=True,
                key="cat_action"
            )
            
            # ===== NUEVA SECCIÓN: AGREGAR CATEGORÍA =====
            if cat_action == "➕ Agregar Categoría":
                st.markdown("#### ➕ Crear Nueva Categoría")
                
                with st.form("form_new_category", clear_on_submit=True):
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        new_cat_name = st.text_input(
                            "Nombre de la categoría *",
                            placeholder="ej: Cabello, Uñas, Masaje",
                            max_chars=100
                        )
                        new_cat_icon = st.selectbox(
                            "Ícono 🎨",
                            ["💅", "💇", "💆", "✨", "🎨", "🪮", "📋", "💄", "🧖", "💆‍♀️", "👄", "🏖️"],
                            index=0
                        )
                    
                    with col2:
                        new_cat_desc = st.text_area(
                            "Descripción",
                            placeholder="Describe brevemente esta categoría de servicios",
                            max_chars=200
                        )
                        new_cat_color = st.color_picker("Color de la categoría", value="#EC4899")
                    
                    # Botón para crear
                    submitted = st.form_submit_button("✅ Crear Categoría", use_container_width=True)
                    
                    if submitted:
                        # Validar que no esté vacío
                        if not new_cat_name.strip():
                            st.error("❌ El nombre de la categoría no puede estar vacío")
                        else:
                            # Intentar crear
                            success, msg, cat_id = db.create_category(
                                name=new_cat_name.strip(),
                                description=new_cat_desc.strip(),
                                icon=new_cat_icon,
                                color=new_cat_color
                            )
                            
                            if success:
                                st.success(f"✅ {msg}")
                                st.info(f"ID de la categoría: {cat_id}")
                                st.balloons()
                                st.rerun()
                            else:
                                st.error(f"❌ {msg}")
            
            # ===== VER CATEGORÍAS =====
            elif cat_action == "📊 Ver Categorías":
                categories = db.get_active_categories()
                
                if categories:
                    df_cats = pd.DataFrame([
                        {
                            'Icon': cat['icon'],
                            'Nombre': cat['name'],
                            'Servicios': cat['service_count'],
                            'Descripción': cat.get('description', '')
                        }
                        for cat in categories
                    ])
                    
                    st.dataframe(df_cats, use_container_width=True)
                else:
                    st.info("ℹ️ No hay categorías")
            
            # ===== EDITAR CATEGORÍA =====
            elif cat_action == "✏️ Editar Categoría":
                categories = db.get_active_categories()
                
                if categories:
                    selected_cat_edit = st.selectbox(
                        "Selecciona categoría",
                        options=[cat['name'] for cat in categories],
                        key="cat_edit_select"
                    )
                    
                    selected_cat = [c for c in categories if c['name'] == selected_cat_edit][0]
                    
                    st.markdown("---")
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        edit_cat_name = st.text_input("Nombre", value=selected_cat['name'])
                        edit_cat_icon = st.selectbox("Ícono", ["💅", "💇", "💆", "✨", "🎨", "🪮", "📋", "💄"], 
                                                    index=["💅", "💇", "💆", "✨", "🎨", "🪮", "📋", "💄"].index(selected_cat['icon']))
                    
                    with col2:
                        edit_cat_desc = st.text_area("Descripción", value=selected_cat.get('description', ''))
                        edit_cat_color = st.color_picker("Color", value=selected_cat.get('color', '#EC4899'))
                    
                    if st.button("✅ Actualizar Categoría", use_container_width=True):
                        success, msg = db.update_category(
                            selected_cat['id'],
                            name=edit_cat_name,
                            description=edit_cat_desc,
                            icon=edit_cat_icon,
                            color=edit_cat_color
                        )
                        
                        if success:
                            st.success(msg)
                            st.rerun()
                        else:
                            st.error(msg)
            
            # ===== DETECTAR DUPLICADAS =====
            elif cat_action == "🔍 Detectar Duplicadas":
                st.markdown("---")
                
                duplicates = db.get_duplicate_categories()
                
                if duplicates:
                    st.warning(f"⚠️ Se encontraron {len(duplicates)} categorías con variaciones:")
                    
                    for dup in duplicates:
                        st.markdown(f"""
                        **Nombre limpio:** `{dup['clean_name']}`
                        - Total servicios: {dup['total_services']}
                        - Variaciones encontradas: {dup['num_variations']}
                        - Ejemplos: {', '.join(dup['variations'][:3])}
                        """)
                    
                    st.info("Para normalizar manualmente, edita cada servicio y asigna la categoría correcta")
                else:
                    st.success("✅ No se encontraron categorías duplicadas")

    
    # ===== TAB 3: PROFESIONAL-SERVICIO =====
    with tab3:
        st.markdown("### 🔗 Vincular Profesionales con Servicios")
        st.markdown("Asigna qué servicios puede ofrecer cada profesional")
        
        # Obtener profesionales
        with db.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT id, name FROM professionals ORDER BY name")
            professionals = [dict(zip([desc[0] for desc in cursor.description], row)) for row in cursor.fetchall()]
        
        if not professionals:
            st.error("❌ No hay profesionales. Crea uno primero en la pestaña Profesionales.")
        else:
            selected_prof = st.selectbox(
                "Selecciona un Profesional",
                options=[p['id'] for p in professionals],
                format_func=lambda pid: next(p['name'] for p in professionals if p['id'] == pid),
                key="select_prof_services"
            )
            
            # Obtener servicios disponibles
            with db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, name FROM services ORDER BY name")
                all_services = [dict(zip([desc[0] for desc in cursor.description], row)) for row in cursor.fetchall()]
            
            if not all_services:
                st.error("❌ No hay servicios. Crea uno primero en la pestaña Servicios.")
            else:
                # Obtener servicios asignados al profesional
                with db.get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute('''
                        SELECT service_id FROM professional_services
                        WHERE professional_id = %s
                    ''', (selected_prof,))
                    rows = cursor.fetchall()
                    assigned_services = [row[0] if isinstance(row, tuple) else row['service_id'] for row in rows]
                
                prof_name = next(p['name'] for p in professionals if p['id'] == selected_prof)
                st.markdown(f"#### Servicios asignados a {prof_name}")
                
                # Mostrar servicios en forma de checkboxes
                col1, col2 = st.columns(2)
                
                for idx, svc in enumerate(all_services):
                    col = col1 if idx % 2 == 0 else col2
                    with col:
                        is_assigned = svc['id'] in assigned_services
                        new_state = st.checkbox(
                            svc['name'],
                            value=is_assigned,
                            key=f"service_{selected_prof}_{svc['id']}"
                        )
                        
                        # Si cambió el estado
                        if new_state != is_assigned:
                            with db.get_connection() as conn:
                                cursor = conn.cursor()
                                if new_state:
                                    # Agregar servicio
                                    cursor.execute('''
                                        INSERT INTO professional_services (professional_id, service_id)
                                        VALUES (%s, %s)
                                    ''', (selected_prof, svc['id']))
                                    st.success(f"✅ {svc['name']} asignado")
                                else:
                                    # Remover servicio
                                    cursor.execute('''
                                        DELETE FROM professional_services
                                        WHERE professional_id = %s AND service_id = %s
                                    ''', (selected_prof, svc['id']))
                                    st.warning(f"❌ {svc['name']} removido")
                                conn.commit()
                
                st.markdown("---")
                
                # Mostrar matriz de asignaciones
                st.markdown("#### 📊 Matriz de Asignaciones")
                
                with db.get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute('''
                        SELECT p.id, p.name, COUNT(ps.service_id) as service_count
                        FROM professionals p
                        LEFT JOIN professional_services ps ON p.id = ps.professional_id
                        GROUP BY p.id
                        ORDER BY p.name
                    ''')
                    matrix_data = [dict(zip([desc[0] for desc in cursor.description], row)) for row in cursor.fetchall()]
                
                if matrix_data:
                    matrix_df = pd.DataFrame(matrix_data)
                    st.dataframe(
                        matrix_df.rename(columns={'id': 'ID', 'name': 'Profesional', 'service_count': 'Servicios'}),
                        width='stretch',
                        hide_index=True
                    )
    
    # ===== TAB 4: HORARIOS =====
    # ===== REEMPLAZAR TAB 4 (HORARIOS) EN CONFIGURACIÓN =====
# Busca: with tab4:
# Reemplaza SOLO el contenido de tab4 (hasta antes de with tab5:)

    with tab4:
        st.markdown("### ⏰ Gestión de Horarios")
        
        # Subtabs para Crear/Ver horarios
        horario_mode = st.radio(
            "Modo de Horarios",
            ["📅 Crear Horarios Masivos", "👁️ Ver Horarios", "📊 Estadísticas"],
            horizontal=True
        )
        
        # ===== MODO 1: CREAR HORARIOS MASIVOS =====
        if horario_mode == "📅 Crear Horarios Masivos":
            st.markdown("#### ➕ Crear Horarios para un Profesional")
            st.markdown("Genera bloques de disponibilidad para asignar citas")
            
            col1, col2 = st.columns(2)
            
            with col1:
                # Seleccionar profesional
                with db.get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute("SELECT id, name FROM professionals WHERE active = TRUE ORDER BY name")
                    professionals = [dict(zip([desc[0] for desc in cursor.description], row)) for row in cursor.fetchall()]
                
                if professionals:
                    prof_options = {p['name']: p['id'] for p in professionals}
                    selected_prof_name = st.selectbox(
                        "📌 Selecciona Profesional",
                        list(prof_options.keys()),
                        key="schedule_prof_select"
                    )
                    selected_prof_id = prof_options[selected_prof_name]
                else:
                    st.error("❌ No hay profesionales. Crea uno primero.")
                    selected_prof_id = None
            
            with col2:
                st.markdown("")  # Espaciador
            
            if selected_prof_id:
                st.markdown("---")
                
                # Parámetros para crear horarios
                st.markdown("#### ⚙️ Configuración de Horarios")
                
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    start_date = st.date_input(
                        "📅 Fecha Inicio",
                        value=datetime.now().date(),
                        key="schedule_start_date"
                    )
                
                with col2:
                    end_date = st.date_input(
                        "📅 Fecha Fin",
                        value=datetime.now().date() + timedelta(days=30),
                        key="schedule_end_date"
                    )
                
                with col3:
                    st.markdown("")  # Espaciador
                
                st.markdown("---")
                
                # Horarios
                col1, col2 = st.columns(2)
                
                with col1:
                    start_time = st.time_input(
                        "🕐 Hora Inicio",
                        value=datetime.strptime("08:00", "%H:%M").time(),
                        key="schedule_start_time"
                    )
                
                with col2:
                    end_time = st.time_input(
                        "🕐 Hora Fin",
                        value=datetime.strptime("18:00", "%H:%M").time(),
                        key="schedule_end_time"
                    )
                
                st.markdown("---")
                
                # Seleccionar días de la semana
                st.markdown("#### 📆 Días de la Semana")
                
                col1, col2, col3, col4 = st.columns(4)
                
                days_names = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
                days_selected = []
                
                for idx, day_name in enumerate(days_names):
                    if idx < 4:
                        col = [col1, col2, col3, col4][idx]
                    else:
                        col = [col1, col2, col3][idx - 4]
                    
                    with col:
                        if st.checkbox(day_name, value=(idx < 5), key=f"day_{idx}"):
                            days_selected.append(idx)
                
                st.markdown("---")
                
                # Botón para crear horarios
                if st.button("✅ Crear Horarios", width='stretch', key="create_schedules_btn"):
                    if not days_selected:
                        st.error("❌ Selecciona al menos un día de la semana")
                    elif start_date > end_date:
                        st.error("❌ La fecha inicio debe ser menor a la fecha fin")
                    else:
                        success, message = db.create_professional_schedules(
                            professional_id=selected_prof_id,
                            start_date=start_date.strftime('%Y-%m-%d'),
                            end_date=end_date.strftime('%Y-%m-%d'),
                            start_time=start_time.strftime('%H:%M'),
                            end_time=end_time.strftime('%H:%M'),
                            days_of_week=days_selected
                        )
                        
                        if success:
                            st.success(message)
                            st.balloons()
                        else:
                            st.error(message)
        
        # ===== MODO 2: VER HORARIOS =====
        elif horario_mode == "👁️ Ver Horarios":
            st.markdown("#### 📋 Horarios por Profesional")
            
            with db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, name FROM professionals ORDER BY name")
                professionals = [dict(zip([desc[0] for desc in cursor.description], row)) for row in cursor.fetchall()]
            
            if professionals:
                # Seleccionar profesional
                prof_options = {p['name']: p['id'] for p in professionals}
                selected_prof_name = st.selectbox(
                    "📌 Profesional",
                    list(prof_options.keys()),
                    key="view_schedule_prof"
                )
                selected_prof_id = prof_options[selected_prof_name]
                
                # Rango de fechas para visualizar
                col1, col2 = st.columns(2)
                with col1:
                    view_start_date = st.date_input(
                        "Desde",
                        value=datetime.now().date(),
                        key="view_start_date"
                    )
                with col2:
                    view_end_date = st.date_input(
                        "Hasta",
                        value=datetime.now().date() + timedelta(days=30),
                        key="view_end_date"
                    )
                
                # Obtener horarios
                schedules = db.get_professional_schedules(
                    professional_id=selected_prof_id,
                    start_date=view_start_date.strftime('%Y-%m-%d'),
                    end_date=view_end_date.strftime('%Y-%m-%d')
                )
                
                if schedules:
                    st.markdown("---")
                    
                    # Agrupar por fecha
                    from itertools import groupby
                    schedules_by_date = {}
                    for sched in schedules:
                        date = sched['date']
                        if date not in schedules_by_date:
                            schedules_by_date[date] = []
                        schedules_by_date[date].append(sched)
                    
                    # Mostrar horarios agrupados
                    for date in sorted(schedules_by_date.keys()):
                        day_schedules = schedules_by_date[date]
                        
                        # Convertir fecha a nombre del día
                        date_obj = datetime.strptime(date, '%Y-%m-%d')
                        day_name = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo'][date_obj.weekday()]
                        
                        with st.expander(f"📅 {date} ({day_name})"):
                            col1, col2, col3, col4 = st.columns([2, 2, 2, 1])
                            
                            with col1:
                                st.markdown("**Hora**")
                            with col2:
                                st.markdown("**Estado**")
                            with col3:
                                st.markdown("**Acciones**")
                            
                            st.markdown("---")
                            
                            for sched in day_schedules:
                                col1, col2, col3, col4 = st.columns([2, 2, 2, 1])
                                
                                with col1:
                                    st.text(sched['start_time'])
                                
                                with col2:
                                    status = "✅ Disponible" if sched['available'] else "❌ Ocupado"
                                    st.text(status)
                                
                                with col3:
                                    if sched['available']:
                                        if st.button("🗑️ Eliminar", key=f"del_sched_{sched['id']}"):
                                            with db.get_connection() as conn:
                                                cursor = conn.cursor()
                                                cursor.execute("DELETE FROM schedules WHERE id = %s", (sched['id'],))
                                                conn.commit()
                                            st.success("✅ Horario eliminado")
                                            st.rerun()
                                
                                with col4:
                                    st.text("")
                    
                    # Botón para eliminar todos los horarios de este período
                    st.markdown("---")
                    if st.button("🗑️ Eliminar todos los horarios de este período", width='stretch'):
                        success, message = db.delete_professional_schedules(
                            professional_id=selected_prof_id,
                            start_date=view_start_date.strftime('%Y-%m-%d'),
                            end_date=view_end_date.strftime('%Y-%m-%d')
                        )
                        if success:
                            st.success(message)
                            st.rerun()
                        else:
                            st.error(message)
                else:
                    st.info("No hay horarios registrados para este período")
            else:
                st.error("No hay profesionales registrados")
        
        # ===== MODO 3: ESTADÍSTICAS =====
        elif horario_mode == "📊 Estadísticas":
            st.markdown("#### 📊 Estadísticas de Disponibilidad")
            
            with db.get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT id, name FROM professionals ORDER BY name")
                professionals = [dict(zip([desc[0] for desc in cursor.description], row)) for row in cursor.fetchall()]
            
            if professionals:
                # Seleccionar profesional
                prof_options = {p['name']: p['id'] for p in professionals}
                selected_prof_name = st.selectbox(
                    "📌 Profesional",
                    list(prof_options.keys()),
                    key="stats_schedule_prof"
                )
                selected_prof_id = prof_options[selected_prof_name]
                
                # Rango de fechas
                col1, col2 = st.columns(2)
                with col1:
                    stats_start_date = st.date_input(
                        "Desde",
                        value=datetime.now().date(),
                        key="stats_start_date"
                    )
                with col2:
                    stats_end_date = st.date_input(
                        "Hasta",
                        value=datetime.now().date() + timedelta(days=30),
                        key="stats_end_date"
                    )
                
                # Obtener estadísticas
                stats = db.get_schedule_statistics(
                    professional_id=selected_prof_id,
                    start_date=stats_start_date.strftime('%Y-%m-%d'),
                    end_date=stats_end_date.strftime('%Y-%m-%d')
                )
                
                # Mostrar métricas
                st.markdown("---")
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    st.metric("Total Horarios", stats['total'])
                
                with col2:
                    st.metric("Disponibles", stats['available'], delta=f"{stats['utilization_rate']:.1f}% ocupado")
                
                with col3:
                    st.metric("Ocupados", stats['occupied'])
                
                with col4:
                    st.metric("Tasa Utilización", f"{stats['utilization_rate']:.1f}%")
                
                # Gráfico de ocupación
                st.markdown("---")
                st.markdown("#### 📈 Gráfico de Ocupación")
                
                chart_data = {
                    'Estado': ['Disponibles', 'Ocupados'],
                    'Cantidad': [stats['available'], stats['occupied']]
                }
                
                df_chart = pd.DataFrame(chart_data)
                st.bar_chart(df_chart.set_index('Estado'))
            else:
                st.error("No hay profesionales registrados")
    
    # ===== TAB 5: RESPALDO =====
    with tab5:
        st.markdown("### 📋 Respaldo y Mantenimiento")
        
        st.markdown("#### 📊 Estadísticas del Sistema")
        
        with db.get_connection() as conn:
            cursor = conn.cursor()
            
            cursor.execute("SELECT COUNT(*) as count FROM professionals WHERE active = TRUE")
            prof_count = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) as count FROM services WHERE active = TRUE")
            svc_count = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) as count FROM bookings")
            booking_count = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) as count FROM professional_services")
            prof_svc_count = cursor.fetchone()[0]
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Profesionales Activos", prof_count)
        
        with col2:
            st.metric("Servicios Activos", svc_count)
        
        with col3:
            st.metric("Citas Totales", booking_count)
        
        with col4:
            st.metric("Asignaciones Prof-Svc", prof_svc_count)
        
        st.markdown("---")
        st.markdown("#### 💾 Funciones de Respaldo")
        st.info("Funciones de exportación próximamente disponibles")

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: #6B7280; font-size: 0.9rem;'>
    <p>👑 Panel de Administración - Rubí Mata Salón | Hecho con 💖 por tu equipo</p>
</div>
""", unsafe_allow_html=True)